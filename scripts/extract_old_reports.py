#!/usr/bin/env python3
from __future__ import annotations

import argparse
import io
import json
import re
import time
import warnings
import subprocess
from pathlib import Path
from typing import Any

from report_import_utils import (
    REVIEW_COLS,
    clean_spaces,
    format_month,
    guess_from_filename,
    iter_report_files,
    nfc,
    normalize_bool,
    parse_month,
    read_csv_rows,
    resolve_source_file,
    write_csv_rows,
)


DEFAULT_INPUT_DIR = Path("import_reports/2026_03_04")
DEFAULT_INVENTORY = Path("import_reports/report_inventory.csv")
DEFAULT_OUTPUT_CSV = Path("import_reports/extracted_scores_review.csv")
DEFAULT_OUTPUT_XLSX = Path("import_reports/extracted_scores_review.xlsx")
DEFAULT_YEAR = "2026"
DEFAULT_GEMINI_MODEL = "gemini-2.5-flash"

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message="urllib3 v2 only supports OpenSSL")

EXTRACT_FIELDS = [
    "teacher_name",
    "student_name",
    "grade",
    "eval_month",
    "test_name",
    "test_round",
    "score",
    "class_avg",
    "total_students",
    "rank",
    "weak_points",
    "ai_comment",
    "memo",
]


def load_secret(key: str) -> str:
    try:
        import streamlit as st

        value = st.secrets.get(key, "")
        return str(value or "").strip()
    except Exception:
        pass

    secrets_path = Path(".streamlit/secrets.toml")
    if not secrets_path.exists():
        return ""
    text = secrets_path.read_text(encoding="utf-8")
    match = re.search(rf"^\s*{re.escape(key)}\s*=\s*(['\"])(.*?)\1\s*$", text, re.MULTILINE)
    return match.group(2).strip() if match else ""


def configure_gemini(api_key: str):
    if not api_key:
        return None
    try:
        import google.generativeai as genai

        genai.configure(api_key=api_key)
        return genai
    except Exception:
        return None


def gemini_prompt(guessed_name: str, guessed_month: str, source_name: str) -> str:
    return f"""
당신은 학원 성적표 OCR 검수 보조자입니다. 이미지 또는 텍스트에서 성적표 값을 추출하세요.

중요한 확정 정보:
- 파일명 기준 원생명 후보: {guessed_name or "(없음)"}
- 파일명 기준 평가월 후보: {guessed_month or "(없음)"}
- 파일명: {source_name}

규칙:
- 파일명 기준 원생명과 평가월이 있으면 본문 OCR보다 우선합니다.
- 확실하지 않은 값은 추측하지 말고 빈 문자열로 둡니다.
- 시험명은 성적표에 적힌 시험/평가/과목명을 그대로 짧게 적습니다.
- 점수, 반평균, 응시 인원, 등수는 숫자만 적습니다. 없으면 빈 문자열입니다.
- 취약점/보완점, 코멘트, 메모는 본문에 있는 문장만 요약 없이 옮기되 너무 길면 핵심 문장만 남깁니다.
- needs_review는 값이 불확실하거나 핵심 정보가 누락될 때만 true로 둡니다. 충분히 확실하면 false입니다.
- JSON 이외의 설명은 출력하지 마세요.

반드시 아래 JSON 객체 하나만 출력하세요.
{{
  "teacher_name": "",
  "student_name": "",
  "grade": "",
  "eval_month": "",
  "test_name": "",
  "test_round": "",
  "score": "",
  "class_avg": "",
  "total_students": "",
  "rank": "",
  "weak_points": "",
  "ai_comment": "",
  "memo": "",
  "needs_review": false,
  "extraction_note": ""
}}
""".strip()


def parse_json_object(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        data = json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if not match:
            return {}
        try:
            data = json.loads(match.group(0))
        except json.JSONDecodeError:
            return {}
    return data if isinstance(data, dict) else {}


def generate_content_with_retries(model, parts: list[Any], retries: int) -> Any:
    for attempt in range(retries + 1):
        try:
            return model.generate_content(parts)
        except Exception as exc:
            name = type(exc).__name__
            message = str(exc)
            retryable = name in {"ResourceExhausted", "TooManyRequests"} or "429" in message
            if not retryable or attempt >= retries:
                raise
            time.sleep(20 * (attempt + 1))


def extract_with_gemini_text(
    genai,
    model_name: str,
    text: str,
    guessed_name: str,
    guessed_month: str,
    source_name: str,
    retries: int,
) -> dict:
    prompt = gemini_prompt(guessed_name, guessed_month, source_name)
    model = genai.GenerativeModel(model_name)
    response = generate_content_with_retries(model, [prompt, "\n\n성적표 텍스트:\n", text], retries)
    return parse_json_object(getattr(response, "text", "") or "")


def extract_with_gemini_image(
    genai,
    model_name: str,
    image,
    guessed_name: str,
    guessed_month: str,
    source_name: str,
    retries: int,
) -> dict:
    prompt = gemini_prompt(guessed_name, guessed_month, source_name)
    model = genai.GenerativeModel(model_name)
    response = generate_content_with_retries(model, [prompt, image], retries)
    return parse_json_object(getattr(response, "text", "") or "")


def extract_pdf_pages(path: Path) -> list[tuple[int, str, Any | None]]:
    try:
        import fitz
    except Exception:
        return [(1, "", None)]

    pages: list[tuple[int, str, Any | None]] = []
    doc = fitz.open(path)
    for index, page in enumerate(doc, start=1):
        text = clean_spaces(page.get_text("text"))
        pages.append((index, text, None))
    return pages or [(1, "", None)]


def render_pdf_page(path: Path, page_number: int):
    try:
        import fitz
        from PIL import Image
    except Exception:
        return None

    doc = fitz.open(path)
    page = doc.load_page(page_number - 1)
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
    return Image.open(io.BytesIO(pix.tobytes("png")))


def image_for_path(path: Path):
    try:
        from PIL import Image

        return Image.open(path)
    except Exception:
        return None


def heuristic_from_text(text: str) -> dict:
    lines = [clean_spaces(line) for line in text.splitlines() if clean_spaces(line)]
    compact = clean_spaces(text)

    def first_line_match(pattern: str) -> str:
        for line in lines:
            match = re.search(pattern, line)
            if match:
                return clean_spaces(match.group(1) if match.groups() else match.group(0))
        return ""

    def number_near(label: str, occurrence: int = 1, window: int = 4) -> str:
        seen = 0
        for idx, line in enumerate(lines):
            if label not in line:
                continue
            seen += 1
            if seen != occurrence:
                continue
            tail = line.split(label, 1)[-1]
            same_line = re.search(r"\d+(?:\.\d+)?", tail)
            if same_line:
                return same_line.group(0)
            for next_line in lines[idx + 1 : idx + 1 + window]:
                match = re.search(r"\d+(?:\.\d+)?", next_line)
                if match:
                    return match.group(0)
        return ""

    def text_between(start_patterns: list[str], end_patterns: list[str], limit: int = 500) -> str:
        start_idx = -1
        for idx, line in enumerate(lines):
            if any(pattern in line for pattern in start_patterns):
                start_idx = idx
                break
        if start_idx < 0:
            return ""
        end_idx = len(lines)
        for idx in range(start_idx + 1, len(lines)):
            if any(pattern in lines[idx] for pattern in end_patterns):
                end_idx = idx
                break
        return clean_spaces(" ".join(lines[start_idx:end_idx]))[:limit]

    def after(labels: list[str], pattern: str = r"([^\n\r|/]{1,40})") -> str:
        for label in labels:
            match = re.search(rf"{label}\s*[:：]?\s*{pattern}", compact)
            if match:
                return clean_spaces(match.group(1))
        return ""

    def number_after(labels: list[str]) -> str:
        value = after(labels, r"(\d+(?:\.\d+)?)")
        return value

    grade_match = re.search(r"([초중고]\s*\d\s*|[1-6]\s*)학년", compact)
    month_match = re.search(r"(0?[34])\s*월", compact)
    teacher = first_line_match(r"담당\s*[:：]?\s*([^|]{2,20})") or after(["담당교사", "담당 선생님", "교사", "선생님"])
    student = first_line_match(r"([가-힣]{2,5})\s*원생") or after(["원생명", "학생명", "이름"])
    score = number_near("원생 종합") or number_near("원생 점수")
    class_avg = number_near("반 종합 평균") or number_near("반 평균")
    test_name = "월간 종합 평균" if "월간 종합 평균" in compact else after(["시험명", "평가명", "과목", "테스트명"])
    ai_comment = text_between(["이번 달 학습 단원"], ["다음 달에는", "다음 달 로드맵"], 500)
    if not ai_comment:
        ai_match = re.search(r"(종합 평균[^.。]{20,300})", compact)
        ai_comment = clean_spaces(ai_match.group(1)) if ai_match else ""
    memo = ""
    if "담당 강사 확인" in compact:
        before_sign = compact.split("담당 강사 확인", 1)[0]
        memo_match = re.search(r"((?:[1-6]학년|초등|중등|고등).{80,500})$", before_sign)
        if memo_match:
            memo = clean_spaces(memo_match.group(1))
    if not memo:
        memo = text_between(["선생님의 메모"], ["담당 강사 확인"], 500)
    weak = ""
    weak_match = re.search(r"([^.!?。]*?(?:낮아|보완|취약|틀린|실수)[^.!?。]{0,180})", compact)
    if weak_match:
        weak = clean_spaces(weak_match.group(1))

    return {
        "teacher_name": teacher,
        "student_name": student,
        "grade": clean_spaces(grade_match.group(0)) if grade_match else after(["학년"]),
        "eval_month": f"{int(month_match.group(1))}월" if month_match else "",
        "test_name": test_name,
        "test_round": after(["회차", "차시"], r"(\d+)"),
        "score": score or number_after(["원생 점수", "학생 점수", "점수", "종합 점수"]),
        "class_avg": class_avg or number_after(["반 평균", "반평균", "평균"]),
        "total_students": number_after(["응시 인원", "전체 인원", "총원"]),
        "rank": number_after(["등수", "석차"]),
        "weak_points": weak or after(["보완점", "취약점", "약점"], r"(.{1,200})"),
        "ai_comment": ai_comment or after(["종합 코멘트", "AI 코멘트", "코멘트", "총평"], r"(.{1,300})"),
        "memo": memo or after(["메모", "수업 메모"], r"(.{1,200})"),
    }


def local_ocr_text(path: Path) -> str:
    swift_script = Path(__file__).with_name("macos_ocr.swift")
    if not swift_script.exists():
        return ""
    try:
        result = subprocess.run(
            ["swift", str(swift_script), str(path)],
            check=False,
            capture_output=True,
            text=True,
            timeout=60,
        )
    except Exception:
        return ""
    return result.stdout if result.returncode == 0 else ""


def finalize_row(
    extracted: dict,
    guessed_name: str,
    guessed_month: str,
    source_file: str,
    source_page: int,
    inventory_needs_review: bool,
    base_note: str,
) -> dict:
    row = {field: clean_spaces(extracted.get(field, "")) for field in EXTRACT_FIELDS}
    notes = [base_note] if base_note else []
    review_needed = bool(inventory_needs_review)

    model_name = clean_spaces(row.get("student_name"))
    model_month = format_month(row.get("eval_month"))
    if guessed_name:
        if model_name and model_name.replace(" ", "") != guessed_name.replace(" ", ""):
            notes.append(f"본문 원생명({model_name})과 파일명 원생명({guessed_name}) 불일치")
            review_needed = True
        row["student_name"] = guessed_name
    if guessed_month:
        if model_month and parse_month(model_month) != parse_month(guessed_month):
            notes.append(f"본문 평가월({model_month})과 파일명 평가월({guessed_month}) 불일치")
            review_needed = True
        row["eval_month"] = guessed_month

    row["year"] = DEFAULT_YEAR
    row["score"] = re.sub(r"[^0-9.]", "", row.get("score", ""))
    row["class_avg"] = re.sub(r"[^0-9.]", "", row.get("class_avg", ""))
    row["total_students"] = re.sub(r"[^0-9.]", "", row.get("total_students", ""))
    row["rank"] = re.sub(r"[^0-9.]", "", row.get("rank", ""))

    critical = ["student_name", "eval_month", "grade", "test_name", "score", "class_avg", "ai_comment"]
    missing = [field for field in critical if not clean_spaces(row.get(field, ""))]
    if missing:
        notes.append("필수 검수값 누락: " + ", ".join(missing))
        review_needed = True

    model_review = normalize_bool(extracted.get("needs_review", False))
    if model_review:
        review_needed = True
    model_note = clean_spaces(extracted.get("extraction_note", ""))
    if model_note:
        notes.append(model_note)

    row["created_at"] = ""
    row["source_file"] = source_file
    row["source_page"] = str(source_page)
    row["needs_review"] = str(review_needed)
    row["extraction_note"] = "; ".join(dict.fromkeys(note for note in notes if note))
    return {col: row.get(col, "") for col in REVIEW_COLS}


def inventory_rows(input_dir: Path, inventory_path: Path) -> list[dict]:
    if inventory_path.exists():
        return read_csv_rows(inventory_path)
    rows = []
    for path in iter_report_files(input_dir):
        guessed = guess_from_filename(path)
        rows.append(
            {
                "source_file": nfc(str(path)),
                "guessed_student_name": guessed["guessed_student_name"],
                "guessed_month": guessed["guessed_month"],
                "file_ext": guessed["file_ext"],
                "needs_review": str(bool(guessed["needs_review"])),
                "note": guessed["note"],
            }
        )
    return rows


def write_xlsx(path: Path, rows: list[dict]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    ws.append(REVIEW_COLS)
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([row.get(col, "") for col in REVIEW_COLS])
    ws.freeze_panes = "A2"
    for idx, col in enumerate(REVIEW_COLS, start=1):
        max_len = max([len(str(col))] + [len(str(row.get(col, ""))) for row in rows])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 45)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def extract_one_file(path: Path, inv: dict, genai, model_name: str, retries: int) -> list[dict]:
    guessed_name = clean_spaces(inv.get("guessed_student_name", ""))
    guessed_month = format_month(inv.get("guessed_month", ""))
    inventory_review = normalize_bool(inv.get("needs_review", False))
    inventory_note = clean_spaces(inv.get("note", ""))
    source_file = nfc(str(path))

    rows: list[dict] = []
    if path.suffix.lower() == ".pdf":
        pages = extract_pdf_pages(path)
        for page_number, text, _image in pages:
            extracted: dict = {}
            note = inventory_note
            if genai and len(text) >= 80:
                try:
                    extracted = extract_with_gemini_text(genai, model_name, text, guessed_name, guessed_month, path.name, retries)
                    note = "; ".join(filter(None, [note, "PDF 텍스트 레이어 사용"]))
                except Exception as exc:
                    note = "; ".join(filter(None, [note, f"Gemini 텍스트 추출 실패: {type(exc).__name__}"]))
            if not extracted and len(text) >= 80:
                extracted = heuristic_from_text(text)
                note = "; ".join(filter(None, [note, "PDF 텍스트 휴리스틱 사용"]))
            if not extracted and genai:
                image = render_pdf_page(path, page_number)
                if image is not None:
                    try:
                        extracted = extract_with_gemini_image(genai, model_name, image, guessed_name, guessed_month, path.name, retries)
                        note = "; ".join(filter(None, [note, "PDF 페이지 이미지 Gemini 사용"]))
                    except Exception as exc:
                        note = "; ".join(filter(None, [note, f"Gemini 이미지 추출 실패: {type(exc).__name__}"]))
            if not extracted:
                note = "; ".join(filter(None, [note, "추출 실패: 텍스트 레이어 부족 또는 Gemini 사용 불가"]))
            rows.append(finalize_row(extracted, guessed_name, guessed_month, source_file, page_number, inventory_review, note))
        return rows

    extracted = {}
    note = inventory_note
    image = image_for_path(path)
    if image is None:
        note = "; ".join(filter(None, [note, "이미지 파일을 열 수 없음"]))
    elif genai:
        try:
            extracted = extract_with_gemini_image(genai, model_name, image, guessed_name, guessed_month, path.name, retries)
            note = "; ".join(filter(None, [note, "이미지 Gemini 사용"]))
        except Exception as exc:
            note = "; ".join(filter(None, [note, f"Gemini 이미지 추출 실패: {type(exc).__name__}"]))
    else:
        note = "; ".join(filter(None, [note, "Gemini 미사용 또는 사용 불가"]))
    if not extracted:
        ocr_text = local_ocr_text(path)
        if ocr_text:
            extracted = heuristic_from_text(ocr_text)
            note = "; ".join(filter(None, [note, "macOS Vision OCR 휴리스틱 사용"]))

    return [finalize_row(extracted, guessed_name, guessed_month, source_file, 1, inventory_review, note)]


def main() -> int:
    parser = argparse.ArgumentParser(description="3월/4월 출력본 성적표를 검수용 CSV/XLSX로 추출합니다.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR), help="성적표 파일 폴더")
    parser.add_argument("--inventory", default=str(DEFAULT_INVENTORY), help="report_inventory.csv 경로")
    parser.add_argument("--output-csv", default=str(DEFAULT_OUTPUT_CSV), help="검수 CSV 경로")
    parser.add_argument("--output-xlsx", default=str(DEFAULT_OUTPUT_XLSX), help="검수 XLSX 경로")
    parser.add_argument("--limit", type=int, default=0, help="앞에서부터 N개 파일만 처리")
    parser.add_argument("--model", default=DEFAULT_GEMINI_MODEL, help="Gemini 모델명")
    parser.add_argument("--retries", type=int, default=2, help="Gemini 속도 제한/429 재시도 횟수")
    parser.add_argument("--delay", type=float, default=0.0, help="파일 처리 사이 대기 초")
    parser.add_argument("--resume", action="store_true", help="기존 output CSV에서 needs_review=False 행은 재사용")
    parser.add_argument("--no-gemini", action="store_true", help="Gemini 호출 없이 가능한 텍스트 추출만 수행")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        raise SystemExit(f"입력 폴더가 없습니다: {input_dir}")

    api_key = "" if args.no_gemini else load_secret("GEMINI_API_KEY")
    genai = None if args.no_gemini else configure_gemini(api_key)
    if not args.no_gemini and not genai:
        print("Gemini를 사용할 수 없어 텍스트 레이어/휴리스틱만 수행합니다. 키 값은 출력하지 않습니다.")

    inventory = inventory_rows(input_dir, Path(args.inventory))
    if args.limit and args.limit > 0:
        inventory = inventory[: args.limit]

    reusable: dict[str, list[dict]] = {}
    if args.resume and Path(args.output_csv).exists():
        for row in read_csv_rows(args.output_csv):
            if not normalize_bool(row.get("needs_review")):
                reusable.setdefault(nfc(Path(row.get("source_file", "")).name), []).append(row)

    rows: list[dict] = []
    for index, inv in enumerate(inventory, start=1):
        path = resolve_source_file(inv.get("source_file", ""), input_dir)
        reusable_rows = reusable.get(nfc(path.name if path else Path(inv.get("source_file", "")).name), [])
        if reusable_rows:
            rows.extend(reusable_rows)
            print(f"[{index}/{len(inventory)}] reusing {nfc(path.name)}", flush=True)
            continue
        if not path.exists():
            guessed = guess_from_filename(inv.get("source_file", ""))
            row = finalize_row(
                {},
                clean_spaces(inv.get("guessed_student_name") or guessed["guessed_student_name"]),
                format_month(inv.get("guessed_month") or guessed["guessed_month"]),
                nfc(inv.get("source_file", "")),
                1,
                True,
                "원본 파일을 찾지 못함",
            )
            rows.append(row)
            continue
        print(f"[{index}/{len(inventory)}] extracting {nfc(path.name)}", flush=True)
        rows.extend(extract_one_file(path, inv, genai, args.model, args.retries))
        if args.delay > 0:
            time.sleep(args.delay)

    write_csv_rows(args.output_csv, rows, REVIEW_COLS)
    xlsx_ok = write_xlsx(Path(args.output_xlsx), rows)
    review_count = sum(1 for row in rows if normalize_bool(row.get("needs_review")))
    print(f"saved {args.output_csv} ({len(rows)} rows, needs_review={review_count})")
    if xlsx_ok:
        print(f"saved {args.output_xlsx}")
    else:
        print("XLSX 저장은 건너뜀: openpyxl 패키지가 없습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
